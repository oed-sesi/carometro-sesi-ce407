#!/usr/bin/env python3
"""
gerenciar_usuarios.py
Carômetro Escolar — SESI 407  |  v5.3

Gerencia os usuários de acesso em data/config.json.

Uso (modo interativo):
    python3 scripts/gerenciar_usuarios.py

Uso (modo direto):
    python3 scripts/gerenciar_usuarios.py --listar
    python3 scripts/gerenciar_usuarios.py --adicionar
    python3 scripts/gerenciar_usuarios.py --alterar-senha  <usuario>
    python3 scripts/gerenciar_usuarios.py --remover        <usuario>
    python3 scripts/gerenciar_usuarios.py --verificar-senha <usuario>
    python3 scripts/gerenciar_usuarios.py --ativar-login
    python3 scripts/gerenciar_usuarios.py --desativar-login
    python3 scripts/gerenciar_usuarios.py --config-escola

Importação em lote via planilha:
    python3 scripts/gerenciar_usuarios.py --importar-lista lista_usuarios.xlsx
    python3 scripts/gerenciar_usuarios.py --importar-lista lista_usuarios.xlsx --dry-run
    python3 scripts/gerenciar_usuarios.py --importar-lista lista_usuarios.xlsx --forcar-senha
    python3 scripts/gerenciar_usuarios.py --exportar-lista lista_usuarios_export.xlsx

Formato da planilha lista_usuarios.xlsx:
    numero_nif, nome_colaborador, funcao, tipo_acesso, senha_acesso
    - usuario   : numero_nif
    - senha     : <numero_nif>@sesi407  (gerada automaticamente)
    - tipo_acesso: admin | visualizador (mapeado para admin | viewer)
"""

from __future__ import annotations

import getpass
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# ── Instalação silenciosa de dependências ──────────────────────
import importlib.util

def _ensure_deps() -> None:
    import subprocess
    for pkg, mod in [("openpyxl", "openpyxl"), ("pandas", "pandas")]:
        if importlib.util.find_spec(mod) is None:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pkg,
                 "--break-system-packages", "-q"]
            )

# ── Caminhos ───────────────────────────────────────────────────
BASE        = Path(__file__).resolve().parent.parent
CONFIG_PATH = BASE / "data" / "config.json"

SEP  = "─" * 62
SEP2 = "═" * 62

# ── Config padrão ──────────────────────────────────────────────
CONFIG_DEFAULT: dict = {
    "app": {
        "nome":       "Carômetro Escolar",
        "subtitulo":  "Centro Educacional SESI 407",
        "escola":     "SESI 407",
        "cidade":     "São José do Rio Preto — SP",
        "ano_letivo": str(datetime.now().year),
    },
    "auth": {
        "habilitado": True,
        "usuarios":   [],
    },
    "series_ordem": [
        "6º Ano", "7º Ano", "8º Ano", "9º Ano",
        "1º EM", "2º EM", "3º EM",
    ],
}

# ── Mapeamento de tipo_acesso → perfil ─────────────────────────
MAPA_PERFIL = {
    "admin":        "admin",
    "administrador":"admin",
    "visualizador": "viewer",
    "viewer":       "viewer",
    "professor":    "viewer",
    "coordenador":  "viewer",
    "diretor":      "viewer",
}

# ── Mapa função → tratamento de saudação ──────────────────────
FUNCAO_SAUDACAO: dict[str, str] = {
    # ── Sigla ─────────────────────────────────────────────────
    "orientador de educação digital":   "OED",
    "orientadora de educação digital":  "OED",

    # ── Orientação Educacional — título completo ──────────────
    "orientadora educacional":          "Orientadora Educacional",
    "orientador educacional":           "Orientador Educacional",

    # ── Coordenação — curta (EF / EM) já vem com gênero ──────
    "coordenadora ef":  "Coordenadora",
    "coordenador ef":   "Coordenador",
    "coordenadora em":  "Coordenadora",
    "coordenador em":   "Coordenador",
    # Formas longas (legado / variações futuras)
    "coordenadora ensino fundamental":  "Coordenadora",
    "coordenador ensino fundamental":   "Coordenador",
    "coordenado ensino fundamental":    "Coordenador",
    "coordenadora ensino médio":        "Coordenadora",
    "coordenador ensino médio":         "Coordenador",
    "coordenado ensino médio":          "Coordenador",

    # ── Professor / Professora e variações ───────────────────
    "professora":                         "Professora",
    "professor":                          "Professor",
    "professora educaçação inclusiva":    "Professora",
    "professora educação inclusiva":      "Professora",
    "professor educação inclusiva":       "Professor",
    "professora tutora":                  "Professora",
    "professor tutor":                    "Professor",

    # ── Demais funções ────────────────────────────────────────
    "diretora":           "Diretora",
    "diretor":            "Diretor",
    "inspetora":          "Inspetora",
    "inspetor":           "Inspetor",
    "nutricionista":      "Nutricionista",
    "cozinheira":         "Cozinheira",
    "cozinheiro":         "Cozinheiro",
    "auxiliar doscente":  "Auxiliar",
    "auxiliar docente":   "Auxiliar",
    "auxiliar de cozinha":"Auxiliar",
    "secretária":         "Secretária",
    "secretário":         "Secretário",
    "bibliotecária":      "Bibliotecária",
    "bibliotecário":      "Bibliotecário",

    # ── Sem tratamento antes do nome ─────────────────────────
    "manutenção":   "",
    "manutencao":   "",
}

_PARTICULAS = {"de", "da", "do", "dos", "das", "e"}

def primeiro_nome(nome_completo: str) -> str:
    """Retorna o primeiro nome real (ignora partículas)."""
    for parte in nome_completo.strip().split():
        if parte.lower() not in _PARTICULAS:
            return parte
    return nome_completo.strip().split()[0] if nome_completo.strip() else ""

def gerar_saudacao(nome_completo: str, funcao: str, perfil: str) -> str:
    """Gera 'Olá, <tratamento> <PrimeiroNome>' conforme o mapa de funções."""
    pnome = primeiro_nome(nome_completo)

    if not funcao and perfil == "admin":
        return f"Olá, Administrador"

    chave     = (funcao or "").lower().strip()
    tratamento = FUNCAO_SAUDACAO.get(chave)   # None = desconhecida

    if tratamento is None or chave in ("", "atualizar..."):
        return f"Olá, {pnome}"
    if tratamento == "":
        return f"Olá, {pnome}"
    return f"Olá, {tratamento} {pnome}"

# ── Hash ───────────────────────────────────────────────────────
def sha256(texto: str) -> str:
    return hashlib.sha256(texto.encode("utf-8")).hexdigest()

def gerar_senha_nif(nif: str) -> str:
    """Padrão: <numero_nif>@sesi407"""
    return f"{nif.strip()}@sesi407"

# ── Carregar / salvar config ───────────────────────────────────
def carregar_config() -> dict:
    if CONFIG_PATH.is_file():
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    return json.loads(json.dumps(CONFIG_DEFAULT))

def salvar_config(cfg: dict) -> None:
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def usuarios(cfg: dict) -> list[dict]:
    return cfg.setdefault("auth", {}).setdefault("usuarios", [])

# ── Validações ─────────────────────────────────────────────────
def validar_usuario(nome: str) -> str | None:
    if not nome:
        return "Nome de usuário não pode ser vazio."
    if len(nome) < 3:
        return "Nome de usuário deve ter pelo menos 3 caracteres."
    if not re.fullmatch(r"[a-z0-9_.@\-]+", nome.lower()):
        return "Use apenas letras, números, '.', '_', '-', '@'."
    return None

def validar_senha(senha: str) -> str | None:
    if len(senha) < 6:
        return "A senha deve ter pelo menos 6 caracteres."
    return None

def encontrar_usuario(cfg: dict, nome: str) -> dict | None:
    return next(
        (u for u in usuarios(cfg) if u["usuario"].lower() == nome.lower()),
        None
    )

# ── Exibição ───────────────────────────────────────────────────
def cabecalho(titulo: str) -> None:
    print(f"\n{SEP2}")
    print(f"  Carômetro Escolar — {titulo}")
    print(SEP2)

def print_usuarios(cfg: dict, filtro_perfil: str | None = None) -> None:
    lista = usuarios(cfg)
    if filtro_perfil:
        lista = [u for u in lista if u.get("perfil") == filtro_perfil]

    auth = cfg.get("auth", {})
    login_status = "✅  Ativado" if auth.get("habilitado", True) else "❌  Desativado"

    cabecalho("Gerenciamento de Usuários")
    print(f"  Login   : {login_status}")
    print(f"  Total   : {len(usuarios(cfg))} usuário(s)  "
          f"({sum(1 for u in usuarios(cfg) if u.get('perfil')=='admin')} admin, "
          f"{sum(1 for u in usuarios(cfg) if u.get('perfil')!='admin')} viewer)")
    if filtro_perfil:
        print(f"  Filtro  : {filtro_perfil}")
    print()

    if not lista:
        print("  ℹ️   Nenhum usuário encontrado.")
        return

    print(f"  {'#':<4} {'NIF/Usuário':<14} {'Nome':<32} {'Função':<26} {'Perfil':<8} {'Saudação'}")
    print(f"  {SEP}")
    for i, u in enumerate(lista, 1):
        icone    = "🔴" if u.get("perfil") == "admin" else "🔵"
        funcao   = u.get("funcao", "—")[:25]
        nome_exib= u.get("nome", "—")[:31]
        saudacao = gerar_saudacao(u.get("nome",""), u.get("funcao",""), u.get("perfil","viewer"))
        print(f"  {i:<4} {u['usuario']:<14} {nome_exib:<32} {funcao:<26} {icone} {u.get('perfil','viewer'):<8} {saudacao}")
    print()

# ═══════════════════════════════════════════════════════════════
# IMPORTAÇÃO EM LOTE VIA PLANILHA
# ═══════════════════════════════════════════════════════════════
def normalizar_str(s) -> str:
    if s is None:
        return ""
    v = str(s).strip()
    return "" if v.lower() in ("nan", "none", "") else v

def acao_importar_lista(
    caminho_arg: str | None,
    dry_run: bool = False,
    forcar_senha: bool = False,
) -> None:
    _ensure_deps()
    import pandas as pd

    cabecalho("Importação em Lote — lista_usuarios")

    # ── Resolver caminho do arquivo ────────────────────────────
    if caminho_arg:
        caminho = Path(caminho_arg)
        if not caminho.is_absolute():
            # Tentar relativo ao CWD, depois à raiz do projeto
            if not caminho.exists():
                caminho = BASE / caminho_arg
    else:
        # Procurar automaticamente na raiz do projeto
        for nome in ["lista_usuarios.xlsx", "lista_usuarios.csv"]:
            p = BASE / nome
            if p.exists():
                caminho = p
                break
        else:
            print("  ❌  Nenhum arquivo lista_usuarios encontrado.")
            print(f"      Informe o caminho: --importar-lista <arquivo>\n")
            sys.exit(1)

    if not caminho.exists():
        print(f"  ❌  Arquivo não encontrado: {caminho}\n")
        sys.exit(1)

    print(f"  Arquivo  : {caminho.name}")
    print(f"  Modo     : {'🔍 Simulação (dry-run)' if dry_run else '💾 Importação real'}")
    print(f"  Senha    : {'⚠️  Forçar redefinição para todos' if forcar_senha else 'Preservar senha de quem já existe'}")
    print()

    # ── Ler planilha ───────────────────────────────────────────
    ext = caminho.suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm", ".xls"):
            df = pd.read_excel(caminho, dtype=str)
        elif ext == ".csv":
            try:
                df = pd.read_csv(caminho, dtype=str, encoding="utf-8-sig")
            except UnicodeDecodeError:
                df = pd.read_csv(caminho, dtype=str, encoding="latin-1")
        else:
            print(f"  ❌  Formato não suportado: {ext}. Use .xlsx ou .csv\n")
            sys.exit(1)
    except Exception as e:
        print(f"  ❌  Erro ao ler arquivo: {e}\n")
        sys.exit(1)

    # Normalizar nomes das colunas
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # ── Validar colunas obrigatórias ───────────────────────────
    cols_req = ["numero_nif", "nome_colaborador"]
    faltando = [c for c in cols_req if c not in df.columns]
    if faltando:
        print(f"  ❌  Colunas obrigatórias ausentes: {', '.join(faltando)}")
        print(f"      Colunas encontradas: {', '.join(df.columns.tolist())}\n")
        sys.exit(1)

    # ── Carregar config atual ──────────────────────────────────
    cfg = carregar_config()

    # Índice dos existentes: nif → objeto usuário
    existentes_map: dict[str, dict] = {
        u["usuario"]: u for u in usuarios(cfg)
    }

    # ── Processar cada linha ───────────────────────────────────
    novos      : list[dict] = []
    atualizados: list[dict] = []
    ignorados  : list[tuple] = []
    erros      : list[tuple] = []

    linhas_resultado = []  # para exportar de volta ao XLSX

    for idx, row in df.iterrows():
        nif     = normalizar_str(row.get("numero_nif", ""))
        nome    = normalizar_str(row.get("nome_colaborador", ""))
        funcao  = normalizar_str(row.get("funcao", ""))
        tipo_ac = normalizar_str(row.get("tipo_acesso", "viewer")).lower()

        linha_info = {
            "numero_nif":        nif,
            "nome_colaborador":  nome,
            "funcao":            funcao,
            "tipo_acesso":       tipo_ac,
            "senha_acesso":      f"{nif}@sesi407" if nif else "",
            "saudacao_preview":  "",
            "status_cadastro":   "",
            "data_cadastro":     "",
            "observacao":        "",
        }

        # Validações básicas
        if not nif:
            msg = "numero_nif vazio"
            erros.append((idx + 2, nif, nome, msg))
            linha_info["status_cadastro"] = "❌ ERRO"
            linha_info["observacao"]      = msg
            linhas_resultado.append(linha_info)
            continue

        if not nome:
            msg = "nome_colaborador vazio"
            erros.append((idx + 2, nif, nome, msg))
            linha_info["status_cadastro"] = "❌ ERRO"
            linha_info["observacao"]      = msg
            linhas_resultado.append(linha_info)
            continue

        # Validar NIF como usuário
        erro_val = validar_usuario(nif)
        if erro_val:
            erros.append((idx + 2, nif, nome, erro_val))
            linha_info["status_cadastro"] = "❌ ERRO"
            linha_info["observacao"]      = erro_val
            linhas_resultado.append(linha_info)
            continue

        # Mapear tipo de acesso → perfil
        perfil = MAPA_PERFIL.get(tipo_ac, "viewer")

        # Senha padrão
        senha_padrao = gerar_senha_nif(nif)

        # Montar objeto usuário
        obj_usuario: dict = {
            "usuario":    nif,
            "senha_hash": sha256(senha_padrao),
            "nome":       nome,
            "funcao":     funcao if funcao and funcao.lower() != "atualizar..." else "",
            "perfil":     perfil,
        }

        ts_agora = datetime.now().strftime("%d/%m/%Y %H:%M")

        if nif in existentes_map:
            existente = existentes_map[nif]
            existente["nome"]    = obj_usuario["nome"]
            existente["funcao"]  = obj_usuario["funcao"]
            existente["perfil"]  = obj_usuario["perfil"]
            if forcar_senha or not existente.get("senha_hash"):
                existente["senha_hash"] = obj_usuario["senha_hash"]
                obs = "senha redefinida"
            else:
                obs = "senha preservada"
            atualizados.append({"nif": nif, "nome": nome, "perfil": perfil, "obs": obs})
            linha_info["saudacao_preview"] = gerar_saudacao(nome, obj_usuario["funcao"], perfil)
            linha_info["status_cadastro"]  = "🔄 ATUALIZADO"
            linha_info["data_cadastro"]    = ts_agora
            linha_info["observacao"]       = obs
        else:
            if not dry_run:
                usuarios(cfg).append(obj_usuario)
                existentes_map[nif] = obj_usuario
            novos.append({"nif": nif, "nome": nome, "perfil": perfil})
            linha_info["saudacao_preview"] = gerar_saudacao(nome, obj_usuario["funcao"], perfil)
            linha_info["status_cadastro"]  = "✅ CADASTRADO" if not dry_run else "🔍 SIMULADO"
            linha_info["data_cadastro"]    = ts_agora if not dry_run else ""
            linha_info["observacao"]       = f"senha: {senha_padrao}"

        linhas_resultado.append(linha_info)

    # Salvar config
    if not dry_run:
        salvar_config(cfg)

    # ── Relatório ──────────────────────────────────────────────
    total_admin  = sum(1 for u in novos + atualizados
                       if (u.get("perfil") or u.get("perfil","")) == "admin")

    print(f"  {SEP}")
    print(f"  {'Linhas na planilha':<32}: {len(df)}")
    print(f"  {'✅  Novos cadastrados':<32}: {len(novos)}")
    print(f"  {'🔄  Atualizados':<32}: {len(atualizados)}")
    print(f"  {'❌  Erros / ignorados':<32}: {len(erros)}")
    print(f"  {'🔴  Usuários admin':<32}: {sum(1 for u in usuarios(cfg) if u.get('perfil')=='admin')}")
    print(f"  {'🔵  Usuários viewer':<32}: {sum(1 for u in usuarios(cfg) if u.get('perfil')!='admin')}")
    print(f"  {'📋  Total no sistema':<32}: {len(usuarios(cfg))}")
    print(f"  {SEP}")

    if novos:
        print(f"\n  Novos cadastrados ({len(novos)}):")
        for u in novos:
            icone = "🔴" if u["perfil"] == "admin" else "🔵"
            print(f"    {icone} {u['nif']:<12}  {u['nome']}")

    if atualizados:
        print(f"\n  Atualizados ({len(atualizados)}):")
        for u in atualizados:
            print(f"    🔄 {u['nif']:<12}  {u['nome']}  [{u['obs']}]")

    if erros:
        print(f"\n  ⚠️  Erros ({len(erros)}):")
        for linha_num, nif, nome, msg in erros:
            print(f"    Linha {linha_num}: NIF={nif!r} Nome={nome!r} → {msg}")

    if dry_run:
        print(f"\n  ℹ️  Modo simulação — nenhuma alteração salva no config.json")
        print(f"      Remova --dry-run para aplicar.")
    else:
        print(f"\n  💾  config.json atualizado: {CONFIG_PATH}")

    # ── Exportar resultado de volta ao XLSX ────────────────────
    _exportar_resultado_xlsx(caminho, linhas_resultado, dry_run)

    print(f"\n  {SEP}")
    if not dry_run:
        print(f"\n  ✅  Importação concluída!")
        print(f"      Credenciais de acesso: USUÁRIO = NIF  |  SENHA = <NIF>@sesi407")
        print(f"      Exemplo: usuário=1104186  senha=1104186@sesi407\n")
    else:
        print()


def _exportar_resultado_xlsx(
    origem: Path,
    linhas: list[dict],
    dry_run: bool,
) -> None:
    """Salva um novo XLSX com coluna status_cadastro preenchida."""
    _ensure_deps()
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if not linhas:
        return

    sufixo   = "_simulado" if dry_run else "_cadastrado"
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_out = origem.stem + sufixo + f"_{ts}.xlsx"
    saida    = origem.parent / nome_out

    df_out = pd.DataFrame(linhas, columns=[
        "numero_nif", "nome_colaborador", "funcao",
        "tipo_acesso", "senha_acesso", "saudacao_preview",
        "status_cadastro", "data_cadastro", "observacao",
    ])

    df_out.to_excel(saida, index=False, sheet_name="Usuários")

    # Estilizar
    try:
        wb = load_workbook(saida)
        ws = wb.active

        COR_HEADER = "C0392B"
        COR_NOVO   = "D5F5E3"
        COR_ATUALIZ= "D6EAF8"
        COR_ERRO   = "FADBD8"
        COR_SIMUL  = "FEF9E7"

        # Header
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            cell.fill      = PatternFill("solid", start_color=COR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Linhas de dados
        for row in ws.iter_rows(min_row=2):
            status = str(row[5].value or "")
            if "CADASTRADO" in status:
                cor = COR_NOVO
            elif "ATUALIZADO" in status:
                cor = COR_ATUALIZ
            elif "ERRO" in status:
                cor = COR_ERRO
            elif "SIMULADO" in status:
                cor = COR_SIMUL
            else:
                continue
            for cell in row:
                cell.fill = PatternFill("solid", start_color=cor)

        # Larguras
        larguras = [14, 38, 28, 14, 20, 32, 16, 18, 36]
        for i, w in enumerate(larguras, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        wb.save(saida)
    except Exception:
        pass  # Estilização é opcional

    print(f"\n  📄  Relatório exportado: {saida.name}")


# ── Exportar lista atual de usuários ──────────────────────────
def acao_exportar_lista(caminho_saida: str | None) -> None:
    _ensure_deps()
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cabecalho("Exportar Lista de Usuários")
    cfg   = carregar_config()
    lista = usuarios(cfg)

    if not lista:
        print("  ℹ️   Nenhum usuário cadastrado.\n")
        return

    if caminho_saida:
        saida = Path(caminho_saida)
        if not saida.is_absolute():
            saida = BASE / caminho_saida
    else:
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        saida = BASE / f"lista_usuarios_export_{ts}.xlsx"

    linhas = []
    for u in lista:
        linhas.append({
            "numero_nif":       u["usuario"],
            "nome_colaborador": u.get("nome", ""),
            "funcao":           u.get("funcao", ""),
            "tipo_acesso":      u.get("perfil", "viewer"),
            "senha_acesso":     f"{u['usuario']}@sesi407",
            "perfil":           u.get("perfil", "viewer"),
        })

    df = pd.DataFrame(linhas)
    df.to_excel(saida, index=False, sheet_name="Usuários")

    # Estilizar
    try:
        wb = load_workbook(saida)
        ws = wb.active
        for cell in ws[1]:
            cell.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            cell.fill  = PatternFill("solid", start_color="C0392B")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, w in enumerate([14, 38, 28, 14, 20, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(saida)
    except Exception:
        pass

    print(f"  ✅  Exportado: {saida}")
    print(f"  📋  {len(lista)} usuário(s)\n")


# ═══════════════════════════════════════════════════════════════
# AÇÕES INDIVIDUAIS (existentes — mantidas)
# ═══════════════════════════════════════════════════════════════
def acao_listar(filtro: str | None = None) -> None:
    cfg = carregar_config()
    print_usuarios(cfg, filtro_perfil=filtro)

def acao_adicionar() -> None:
    cabecalho("Adicionar Usuário")
    cfg = carregar_config()

    while True:
        nome = input("  Usuário (login / NIF): ").strip().lower()
        erro = validar_usuario(nome)
        if erro:
            print(f"  ❌  {erro}"); continue
        if encontrar_usuario(cfg, nome):
            print(f"  ❌  Usuário '{nome}' já existe."); continue
        break

    nome_exibicao = input("  Nome completo: ").strip()
    if not nome_exibicao:
        nome_exibicao = nome.capitalize()

    funcao = input("  Função/Cargo (opcional): ").strip()

    print("  Perfil:")
    print("    1 → viewer  (visualizador — padrão)")
    print("    2 → admin   (administrador)")
    escolha = input("  Escolha [1/2, padrão=1]: ").strip()
    perfil  = "admin" if escolha == "2" else "viewer"

    print()
    print("  Senha:")
    print(f"    1 → Padrão automática ({nome}@sesi407)")
    print("    2 → Digitar manualmente")
    op_senha = input("  Escolha [1/2, padrão=1]: ").strip()

    if op_senha == "2":
        while True:
            senha = getpass.getpass("  Senha: ")
            erro  = validar_senha(senha)
            if erro:
                print(f"  ❌  {erro}"); continue
            confirma = getpass.getpass("  Confirmar: ")
            if senha != confirma:
                print("  ❌  Senhas não conferem."); continue
            break
    else:
        senha = gerar_senha_nif(nome)
        print(f"  ℹ️   Senha definida: {senha}")

    novo = {
        "usuario":    nome,
        "senha_hash": sha256(senha),
        "nome":       nome_exibicao,
        "funcao":     funcao,
        "perfil":     perfil,
    }
    usuarios(cfg).append(novo)
    salvar_config(cfg)

    icone = "🔴" if perfil == "admin" else "🔵"
    print(f"\n  ✅  Usuário '{nome}' adicionado!")
    print(f"  {icone}  Perfil : {perfil}")
    print(f"      Nome   : {nome_exibicao}")
    if funcao:
        print(f"      Função : {funcao}")
    print()

def acao_alterar_senha(nome_arg: str | None = None) -> None:
    cabecalho("Alterar Senha")
    cfg = carregar_config()

    if not usuarios(cfg):
        print("  ℹ️   Nenhum usuário cadastrado.\n"); return

    if nome_arg:
        nome = nome_arg.lower()
    else:
        print_usuarios(cfg)
        nome = input("  Usuário: ").strip().lower()

    usuario = encontrar_usuario(cfg, nome)
    if not usuario:
        print(f"  ❌  Usuário '{nome}' não encontrado.\n"); return

    print()
    print("  Opções:")
    print(f"    1 → Senha padrão automática ({nome}@sesi407)")
    print("    2 → Digitar nova senha")
    op = input("  Escolha [1/2, padrão=2]: ").strip()

    if op == "1":
        nova = gerar_senha_nif(nome)
        print(f"  ℹ️   Senha definida: {nova}")
    else:
        atual = getpass.getpass("  Senha atual (Enter para pular): ")
        if atual and sha256(atual) != usuario.get("senha_hash", ""):
            print("  ❌  Senha atual incorreta.\n"); return
        while True:
            nova   = getpass.getpass("  Nova senha: ")
            erro   = validar_senha(nova)
            if erro:
                print(f"  ❌  {erro}"); continue
            confirma = getpass.getpass("  Confirmar: ")
            if nova != confirma:
                print("  ❌  Senhas não conferem."); continue
            break

    usuario["senha_hash"] = sha256(nova)
    salvar_config(cfg)
    print(f"\n  ✅  Senha de '{nome}' alterada.\n")

def acao_remover(nome_arg: str | None = None) -> None:
    cabecalho("Remover Usuário")
    cfg = carregar_config()

    if not usuarios(cfg):
        print("  ℹ️   Nenhum usuário cadastrado.\n"); return

    if nome_arg:
        nome = nome_arg.lower()
    else:
        print_usuarios(cfg)
        nome = input("  Usuário a remover: ").strip().lower()

    usuario = encontrar_usuario(cfg, nome)
    if not usuario:
        print(f"  ❌  Usuário '{nome}' não encontrado.\n"); return

    confirmacao = input(f"  Confirmar remoção de '{nome}' ({usuario.get('nome','')})? [s/N]: ").strip().lower()
    if confirmacao not in ("s", "sim", "y"):
        print("  ↩️  Cancelado.\n"); return

    cfg["auth"]["usuarios"] = [u for u in usuarios(cfg) if u["usuario"].lower() != nome]
    salvar_config(cfg)
    print(f"\n  ✅  Usuário '{nome}' removido.\n")

def acao_toggle_login(ativar: bool) -> None:
    cfg = carregar_config()
    cfg.setdefault("auth", {})["habilitado"] = ativar
    salvar_config(cfg)
    estado = "ATIVADO" if ativar else "DESATIVADO"
    icone  = "✅" if ativar else "❌"
    print(f"\n  {icone}  Login {estado} no config.json\n")

def acao_verificar_senha(nome_arg: str | None = None) -> None:
    cabecalho("Verificar Senha")
    cfg = carregar_config()

    if nome_arg:
        nome = nome_arg.lower()
    else:
        nome = input("  Usuário (NIF): ").strip().lower()

    usuario = encontrar_usuario(cfg, nome)
    if not usuario:
        print(f"  ❌  Usuário '{nome}' não encontrado.\n"); return

    senha = getpass.getpass("  Senha a verificar: ")
    if sha256(senha) == usuario.get("senha_hash", ""):
        print(f"\n  ✅  Senha correta para '{nome}'.\n")
    else:
        print(f"\n  ❌  Senha incorreta para '{nome}'.\n")

def acao_editar_app() -> None:
    cabecalho("Configurações da Escola")
    cfg = carregar_config()
    app = cfg.setdefault("app", {})

    campos = [
        ("nome",       "Nome da aplicação",  app.get("nome", "Carômetro Escolar")),
        ("subtitulo",  "Subtítulo",          app.get("subtitulo", "Centro Educacional SESI 407")),
        ("escola",     "Nome da escola",     app.get("escola", "SESI 407")),
        ("cidade",     "Cidade/Estado",      app.get("cidade", "São José do Rio Preto — SP")),
        ("ano_letivo", "Ano letivo",         app.get("ano_letivo", str(datetime.now().year))),
    ]

    print("  Pressione Enter para manter o valor atual.\n")
    for chave, label, atual in campos:
        novo = input(f"  {label} [{atual}]: ").strip()
        app[chave] = novo if novo else atual

    salvar_config(cfg)
    print("\n  ✅  Configurações atualizadas.\n")

def acao_redefinir_todas_senhas() -> None:
    """Redefine todas as senhas para o padrão <nif>@sesi407."""
    cabecalho("Redefinir Todas as Senhas (padrão NIF)")
    cfg = carregar_config()

    lista = usuarios(cfg)
    if not lista:
        print("  ℹ️   Nenhum usuário cadastrado.\n"); return

    print(f"  Isso redefinirá as senhas de {len(lista)} usuário(s)")
    print(f"  Novo padrão: <NIF>@sesi407  (ex: 1104186 → 1104186@sesi407)\n")
    confirmacao = input("  Confirmar? [s/N]: ").strip().lower()
    if confirmacao not in ("s", "sim"):
        print("  ↩️  Cancelado.\n"); return

    for u in lista:
        u["senha_hash"] = sha256(gerar_senha_nif(u["usuario"]))

    salvar_config(cfg)
    print(f"\n  ✅  Senhas de {len(lista)} usuário(s) redefinidas para o padrão NIF.\n")

# ── Menu interativo ────────────────────────────────────────────
def menu_interativo() -> None:
    while True:
        cfg = carregar_config()
        print_usuarios(cfg)
        print("  Opções:")
        print("    1 → Adicionar usuário manualmente")
        print("    2 → Importar planilha lista_usuarios.xlsx")
        print("    3 → Alterar senha de um usuário")
        print("    4 → Redefinir TODAS as senhas para padrão NIF")
        print("    5 → Remover usuário")
        print("    6 → Verificar senha")
        print("    7 → Exportar lista de usuários")
        print("    8 → Ativar/Desativar login")
        print("    9 → Configurações da escola")
        print("    0 → Sair")
        print()
        op = input("  Escolha: ").strip()
        print()

        if   op == "1": acao_adicionar()
        elif op == "2":
            arq = input("  Caminho da planilha [lista_usuarios.xlsx]: ").strip()
            if not arq: arq = "lista_usuarios.xlsx"
            dr  = input("  Dry-run? (simular sem salvar) [s/N]: ").strip().lower()
            frc = input("  Forçar redefinição de senhas? [s/N]: ").strip().lower()
            acao_importar_lista(arq, dry_run=(dr in ("s","sim")),
                                forcar_senha=(frc in ("s","sim")))
        elif op == "3": acao_alterar_senha()
        elif op == "4": acao_redefinir_todas_senhas()
        elif op == "5": acao_remover()
        elif op == "6": acao_verificar_senha()
        elif op == "7": acao_exportar_lista(None)
        elif op == "8":
            atual = cfg.get("auth", {}).get("habilitado", True)
            acao_toggle_login(not atual)
        elif op == "9": acao_editar_app()
        elif op == "0":
            print("  Até logo!\n"); break
        else:
            print("  ❌  Opção inválida.\n")

# ── Entry point ────────────────────────────────────────────────
def main() -> None:
    args = sys.argv[1:]

    if not args:
        menu_interativo()
        return

    cmd  = args[0].lower()
    arg2 = args[1] if len(args) > 1 else None

    # Flags globais
    dry_run      = "--dry-run"      in args
    forcar_senha = "--forcar-senha" in args

    if   cmd in ("--listar",             "-l"):
        filtro = arg2 if arg2 and not arg2.startswith("--") else None
        acao_listar(filtro)
    elif cmd in ("--adicionar",          "-a"):  acao_adicionar()
    elif cmd in ("--alterar-senha",      "-s"):
        u = arg2 if arg2 and not arg2.startswith("--") else None
        acao_alterar_senha(u)
    elif cmd in ("--remover",            "-r"):
        u = arg2 if arg2 and not arg2.startswith("--") else None
        acao_remover(u)
    elif cmd in ("--verificar-senha",    "-v"):
        u = arg2 if arg2 and not arg2.startswith("--") else None
        acao_verificar_senha(u)
    elif cmd in ("--ativar-login",):             acao_toggle_login(True)
    elif cmd in ("--desativar-login",):          acao_toggle_login(False)
    elif cmd in ("--config-escola",      "-c"):  acao_editar_app()
    elif cmd in ("--redefinir-senhas",):         acao_redefinir_todas_senhas()
    elif cmd in ("--importar-lista",     "-i"):
        arq = arg2 if arg2 and not arg2.startswith("--") else None
        acao_importar_lista(arq, dry_run=dry_run, forcar_senha=forcar_senha)
    elif cmd in ("--exportar-lista",     "-e"):
        saida = arg2 if arg2 and not arg2.startswith("--") else None
        acao_exportar_lista(saida)
    else:
        print(f"\n  ❌  Comando desconhecido: {cmd}")
        print("""
  Comandos disponíveis:
    --listar                           Lista todos os usuários
    --listar admin                     Lista apenas admins
    --adicionar                        Adicionar usuário manualmente
    --importar-lista [arquivo.xlsx]    Importar planilha em lote
      --dry-run                          Simular sem salvar
      --forcar-senha                     Redefinir senha de quem já existe
    --exportar-lista [arquivo.xlsx]    Exportar lista atual
    --alterar-senha  [usuario]         Alterar senha
    --redefinir-senhas                 Redefinir TODAS para padrão NIF
    --remover        [usuario]         Remover usuário
    --verificar-senha [usuario]        Verificar se senha está correta
    --ativar-login                     Ativar login na aplicação
    --desativar-login                  Desativar login na aplicação
    --config-escola                    Editar configurações da escola
""")
        sys.exit(1)


if __name__ == "__main__":
    main()
