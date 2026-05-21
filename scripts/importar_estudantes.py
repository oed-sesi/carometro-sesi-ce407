#!/usr/bin/env python3
"""
importar_estudantes.py
Carômetro Escolar — SESI 407

Funções principais:
  1. Sem argumentos      → cria template_estudantes.csv e template_estudantes.xlsx (só cabeçalho)
  2. Com arquivo         → importa dados, gera alunos.json e organiza imagens por turma

Uso:
  python3 scripts/importar_estudantes.py
  python3 scripts/importar_estudantes.py template_estudantes.csv
  python3 scripts/importar_estudantes.py template_estudantes.xlsx
  python3 scripts/importar_estudantes.py --organizar-imagens   (só reorganiza, sem reimportar)

Colunas do template (obrigatórias*):
  rm_estudante*, ra_estudante, nome_estudante*, serie*, turma*, turno*,
  data_nascimento, email_estudante, status, data_ingresso,
  mae_nome, mae_telefone, mae_email, pai_nome, pai_telefone, pai_email,
  endereco_estudante, termo_autorizado, foto_coletada, observacoes

Valores aceitos:
  termo_autorizado / foto_coletada : Sim | Não | sim | não | true | false | 1 | 0
  data_nascimento / data_ingresso  : DD/MM/YYYY ou YYYY-MM-DD
  status                           : Ativo (padrão) | Inativo | Transferido | Formado
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

# ── Instalação silenciosa de dependências ──────────────────────────────────────
import importlib.util

def _ensure_deps():
    import subprocess
    for pkg, mod in [("openpyxl", "openpyxl"), ("pandas", "pandas")]:
        if importlib.util.find_spec(mod) is None:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pkg, "--break-system-packages", "-q"]
            )

_ensure_deps()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Caminhos base ──────────────────────────────────────────────────────────────
BASE        = Path(__file__).resolve().parent.parent
DATA_DIR    = BASE / "data"
IMAGES_DIR  = BASE / "images"
JSON_PATH   = DATA_DIR / "alunos.json"
TMPL_CSV    = BASE / "template_estudantes.csv"
TMPL_XLSX   = BASE / "template_estudantes.xlsx"

# ── Definição de colunas ───────────────────────────────────────────────────────
COLUNAS_TEMPLATE = [
    "rm_estudante",
    "ra_estudante",
    "nome_estudante",
    "serie",
    "turma",
    "turno",
    "data_nascimento",
    "email_estudante",
    "status",
    "data_ingresso",
    "mae_nome",
    "mae_telefone",
    "mae_email",
    "pai_nome",
    "pai_telefone",
    "pai_email",
    "endereco_estudante",
    "termo_autorizado",
    "foto_coletada",
    "observacoes",
]

COLUNAS_OBRIGATORIAS = ["rm_estudante", "nome_estudante", "serie", "turma", "turno"]

DESCRICOES_COLUNAS = {
    "rm_estudante":      "Matrícula/RM do estudante (ex: 006810)",
    "ra_estudante":      "Registro do Aluno SESI (ex: 000120628062-1)",
    "nome_estudante":    "Nome completo (ex: ANA BEATRIZ SILVA SANTOS)",
    "serie":             "Série (ex: 6º Ano | 7º Ano | 8º Ano | 9º Ano | 1º EM | 2º EM | 3º EM)",
    "turma":             "Turma (ex: A | B | C)",
    "turno":             "Turno (ex: Manhã | Tarde | Integral)",
    "data_nascimento":   "Data de nascimento no formato DD/MM/AAAA (ex: 15/03/2012)",
    "email_estudante":   "E-mail do estudante (ex: ana.santos@portalsesisp.org.br)",
    "status":            "Status: Ativo (padrão) | Inativo | Transferido | Formado",
    "data_ingresso":     "Data de ingresso no formato DD/MM/AAAA (ex: 01/02/2024)",
    "mae_nome":          "Nome completo da mãe/responsável 1",
    "mae_telefone":      "Telefone da mãe/responsável 1 (ex: (16) 99999-8888)",
    "mae_email":         "E-mail da mãe/responsável 1",
    "pai_nome":          "Nome completo do pai/responsável 2",
    "pai_telefone":      "Telefone do pai/responsável 2 (ex: (16) 98888-7777)",
    "pai_email":         "E-mail do pai/responsável 2",
    "endereco_estudante":"Endereço completo (ex: Rua das Flores, 123 — SJRP)",
    "termo_autorizado":  "Autorização LGPD: Sim | Não",
    "foto_coletada":     "Foto já coletada: Sim | Não",
    "observacoes":       "Observações gerais sobre o estudante",
}

EXEMPLO_LINHA = {
    "rm_estudante":      "006810",
    "ra_estudante":      "000120628062-1",
    "nome_estudante":    "ANA BEATRIZ SILVA SANTOS",
    "serie":             "6º Ano",
    "turma":             "A",
    "turno":             "Manhã",
    "data_nascimento":   "15/03/2012",
    "email_estudante":   "ana.santos@portalsesisp.org.br",
    "status":            "Ativo",
    "data_ingresso":     "01/02/2024",
    "mae_nome":          "MARIA DA SILVA SANTOS",
    "mae_telefone":      "(16) 99999-8888",
    "mae_email":         "maria.santos@email.com",
    "pai_nome":          "JOSÉ CARLOS SANTOS",
    "pai_telefone":      "(16) 98888-7777",
    "pai_email":         "jose.santos@email.com",
    "endereco_estudante":"Rua das Flores, 123",
    "termo_autorizado":  "Sim",
    "foto_coletada":     "Não",
    "observacoes":       "Estudante com restrição alimentar",
}

# ── Utilitários ────────────────────────────────────────────────────────────────
SEP  = "─" * 60
SEP2 = "═" * 60

def norm_header(h: str) -> str:
    return str(h).strip().lower().replace(" ", "_").replace("-", "_")

def parse_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    return str(val).strip().lower() in ("sim", "s", "true", "1", "yes", "t")

def parse_date(val) -> str:
    if not val or (isinstance(val, float) and str(val) == "nan"):
        return ""
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return ""
    # pandas Timestamp
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    print(f"    ⚠️  Data não reconhecida: '{s}' — mantida como texto")
    return s

def clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none") else s

def slugify_turma(serie: str, turma: str) -> str:
    """
    Converte série e turma para slug de pasta.
    Ex: '6º Ano', 'A'  → 'images_6ano_A'
        '1º EM',  'B'  → 'images_1serie_B'
        '7º Ano', 'C'  → 'images_7ano_C'
    """
    s = serie.strip()
    t = turma.strip().upper()

    # Extrair número
    num_match = re.search(r"(\d+)", s)
    num = num_match.group(1) if num_match else "X"

    s_lower = s.lower()
    if "em" in s_lower or "série" in s_lower or "serie" in s_lower:
        tipo = "serie"
    else:
        tipo = "ano"

    return f"images_{num}{tipo}_{t}"

# ── Criar template ─────────────────────────────────────────────────────────────
def criar_template_csv(destino: Path) -> None:
    import csv
    with open(destino, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUNAS_TEMPLATE)
        writer.writeheader()
        writer.writerow(EXEMPLO_LINHA)
    print(f"  ✅  Template CSV criado : {destino}")

def criar_template_xlsx(destino: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudantes"

    # Estilos
    cor_header  = "C0392B"   # vermelho SESI
    cor_obrig   = "F9EBEA"   # rosa claro para obrigatórias
    cor_opcional= "FDFEFE"   # quase branco
    cor_exemplo = "EBF5FB"   # azul claro

    font_header  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    font_desc    = Font(italic=True, color="7F8C8D", name="Calibri", size=9)
    font_exemplo = Font(color="1A5276", name="Calibri", size=10)
    font_label   = Font(bold=True, color="C0392B", name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BDC3C7")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Linha 1: cabeçalhos
    for col_idx, col in enumerate(COLUNAS_TEMPLATE, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        is_obrig = col in COLUNAS_OBRIGATORIAS
        cell.font  = font_header
        cell.fill  = PatternFill("solid", start_color=cor_header if is_obrig else "7F8C8D")
        cell.alignment = align_center
        cell.border = borda

    # Linha 2: descrições
    for col_idx, col in enumerate(COLUNAS_TEMPLATE, start=1):
        desc = DESCRICOES_COLUNAS.get(col, "")
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font  = font_desc
        cell.fill  = PatternFill("solid", start_color=cor_obrig if col in COLUNAS_OBRIGATORIAS else cor_opcional)
        cell.alignment = align_left
        cell.border = borda

    # Linha 3: exemplo
    for col_idx, col in enumerate(COLUNAS_TEMPLATE, start=1):
        val = EXEMPLO_LINHA.get(col, "")
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font  = font_exemplo
        cell.fill  = PatternFill("solid", start_color=cor_exemplo)
        cell.alignment = align_left
        cell.border = borda

    # Larguras de coluna
    larguras = {
        "rm_estudante": 14, "ra_estudante": 20, "nome_estudante": 35,
        "serie": 12, "turma": 8, "turno": 12,
        "data_nascimento": 18, "email_estudante": 32, "status": 14,
        "data_ingresso": 18, "mae_nome": 30, "mae_telefone": 18,
        "mae_email": 28, "pai_nome": 30, "pai_telefone": 18,
        "pai_email": 28, "endereco_estudante": 38,
        "termo_autorizado": 18, "foto_coletada": 16, "observacoes": 40,
    }
    for col_idx, col in enumerate(COLUNAS_TEMPLATE, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col, 18)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    # Aba de instruções
    wi = wb.create_sheet("Instruções")
    instrucoes = [
        ("CARÔMETRO ESCOLAR — SESI 407", True, "C0392B"),
        ("Template de Importação de Estudantes", True, "2C3E50"),
        ("", False, None),
        ("REGRAS GERAIS:", True, "C0392B"),
        ("• Preencha a partir da linha 4 da aba 'Estudantes' (linhas 1-3 são cabeçalho+exemplo)", False, "2C3E50"),
        ("• NÃO altere os nomes das colunas na linha 1", False, "2C3E50"),
        ("• Campos obrigatórios: rm_estudante, nome_estudante, serie, turma, turno", False, "C0392B"),
        ("", False, None),
        ("FORMATOS:", True, "C0392B"),
        ("• Datas: DD/MM/AAAA  (ex: 15/03/2012)", False, "2C3E50"),
        ("• termo_autorizado / foto_coletada: Sim ou Não", False, "2C3E50"),
        ("• serie: 6º Ano | 7º Ano | 8º Ano | 9º Ano | 1º EM | 2º EM | 3º EM", False, "2C3E50"),
        ("• turno: Manhã | Tarde | Integral", False, "2C3E50"),
        ("• status: Ativo | Inativo | Transferido | Formado", False, "2C3E50"),
        ("", False, None),
        ("FOTOS:", True, "C0392B"),
        ("• As fotos devem ser nomeadas como:  <rm_estudante>.jpg", False, "2C3E50"),
        ("• Exemplo: 006810.jpg", False, "2C3E50"),
        ("• Coloque todas as fotos na pasta  images/  do projeto", False, "2C3E50"),
        ("• O script organiza automaticamente as fotos por turma após a importação", False, "2C3E50"),
    ]
    wi.column_dimensions["A"].width = 75
    for row_idx, (texto, bold, cor) in enumerate(instrucoes, start=1):
        cell = wi.cell(row=row_idx, column=1, value=texto)
        cell.font = Font(
            bold=bold,
            color=cor or "000000",
            name="Calibri",
            size=12 if bold else 11,
        )
        cell.alignment = Alignment(horizontal="left", vertical="center")
        wi.row_dimensions[row_idx].height = 18

    wb.save(destino)
    print(f"  ✅  Template XLSX criado: {destino}")

# ── Ler arquivo (CSV ou XLSX) ──────────────────────────────────────────────────
def ler_arquivo(caminho: Path) -> tuple[list[dict], list[str]]:
    """Retorna (linhas_como_dict_normalizados, headers_normalizados)."""
    ext = caminho.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(
            caminho,
            dtype=str,
            skiprows=1,          # pula linha de descrições
            header=0,
        )
        # O read_excel com skiprows=1 pega a linha de DESCRIÇÕES como header
        # Vamos reler sem skiprows e tratar corretamente
        df = pd.read_excel(caminho, dtype=str, header=0)
        # Se a linha 1 (índice 0) for de cabeçalho e linha 2 de descrições:
        # detectar se segunda linha parece uma descrição
        if len(df) >= 2:
            # Checar se a segunda linha tem texto longo (descrição)
            segunda = df.iloc[0]
            if any(len(str(v)) > 40 for v in segunda if pd.notna(v)):
                df = df.iloc[1:].reset_index(drop=True)  # pula linha de descrição
    elif ext == ".csv":
        try:
            df = pd.read_csv(caminho, dtype=str, encoding="utf-8-sig")
        except UnicodeDecodeError:
            df = pd.read_csv(caminho, dtype=str, encoding="latin-1")
        # Se segunda linha for de descrição, pular
        if len(df) >= 1:
            primeira = df.iloc[0]
            if any(len(str(v)) > 40 for v in primeira if pd.notna(v)):
                df = df.iloc[1:].reset_index(drop=True)
    else:
        raise ValueError(f"Formato não suportado: {ext}. Use .csv ou .xlsx")

    headers = [norm_header(c) for c in df.columns]
    df.columns = headers
    linhas = df.to_dict("records")
    return linhas, headers

# ── Construir objeto aluno ─────────────────────────────────────────────────────
def construir_aluno(row: dict, num_linha: int) -> dict | None:
    rm     = clean(row.get("rm_estudante", ""))
    nome   = clean(row.get("nome_estudante", ""))
    serie  = clean(row.get("serie", ""))
    turma  = clean(row.get("turma", ""))
    turno  = clean(row.get("turno", ""))

    erros = []
    if not rm:   erros.append("rm_estudante vazio")
    if not nome: erros.append("nome_estudante vazio")
    if not serie:erros.append("serie vazio")
    if not turma:erros.append("turma vazio")
    if not turno:erros.append("turno vazio")

    if erros:
        print(f"    ⚠️  Linha {num_linha}: {', '.join(erros)} — ignorada")
        return None

    return {
        "matricula":        rm,
        "rm_estudante":     rm,
        "ra_estudante":     clean(row.get("ra_estudante", "")),
        "nome_completo":    nome,
        "serie":            serie,
        "turma":            turma.upper(),
        "turno":            turno,
        "status":           clean(row.get("status", "")) or "Ativo",
        "data_nascimento":  parse_date(row.get("data_nascimento", "")),
        "data_ingresso":    parse_date(row.get("data_ingresso", "")),
        "email_estudante":  clean(row.get("email_estudante", "")),
        "mae_nome":         clean(row.get("mae_nome", "")),
        "mae_telefone":     clean(row.get("mae_telefone", "")),
        "mae_email":        clean(row.get("mae_email", "")),
        "pai_nome":         clean(row.get("pai_nome", "")),
        "pai_telefone":     clean(row.get("pai_telefone", "")),
        "pai_email":        clean(row.get("pai_email", "")),
        "endereco":         clean(row.get("endereco_estudante", "") or row.get("endereco", "")),
        "observacoes":      clean(row.get("observacoes", "")),
        "termo_autorizado": parse_bool(row.get("termo_autorizado", "")),
        "foto_coletada":    parse_bool(row.get("foto_coletada", "")),
    }

# ── Organizar imagens por turma ────────────────────────────────────────────────
EXTENSOES_IMG = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}

def organizar_imagens(alunos: list[dict], verbose: bool = True) -> dict:
    """
    Para cada aluno com foto_coletada=True:
      1. Procura images/<rm>.jpg (ou .jpeg/.png/.webp)
      2. Copia para images_<slug>/  dentro da raiz do projeto
      3. Mantém o original em images/ (não remove)

    Retorna estatísticas.
    """
    if not IMAGES_DIR.exists():
        if verbose:
            print(f"  ℹ️  Pasta images/ não encontrada: {IMAGES_DIR}")
        return {"copiadas": 0, "nao_encontradas": [], "turmas": {}}

    copiadas       = 0
    nao_encontradas = []
    turmas_criadas  = set()

    # Índice: rm_estudante → aluno
    rm_map = {a["rm_estudante"]: a for a in alunos}

    # Varrer imagens existentes
    for img_file in sorted(IMAGES_DIR.iterdir()):
        if not img_file.is_file():
            continue
        if img_file.suffix.lower() not in EXTENSOES_IMG:
            continue

        rm = img_file.stem  # ex: "006810"
        aluno = rm_map.get(rm)
        if not aluno:
            continue  # imagem sem aluno correspondente — ignorar

        slug   = slugify_turma(aluno["serie"], aluno["turma"])
        pasta  = BASE / slug
        pasta.mkdir(parents=True, exist_ok=True)
        turmas_criadas.add(slug)

        destino = pasta / f"{rm}.jpg"
        shutil.copy2(img_file, destino)
        copiadas += 1

    # Alunos com foto_coletada=True mas imagem não encontrada
    for aluno in alunos:
        if not aluno["foto_coletada"]:
            continue
        rm = aluno["rm_estudante"]
        encontrou = any(
            (IMAGES_DIR / f"{rm}{ext}").exists()
            for ext in EXTENSOES_IMG
        )
        if not encontrou:
            nao_encontradas.append(rm)

    return {
        "copiadas":        copiadas,
        "nao_encontradas": nao_encontradas,
        "turmas":          sorted(turmas_criadas),
    }

# ── Importar arquivo e gerar JSON ─────────────────────────────────────────────
def importar(caminho: Path) -> None:
    print(f"\n{SEP2}")
    print("  Carômetro Escolar — Importação de Estudantes")
    print(SEP2)
    print(f"  Arquivo : {caminho.name}")
    print(f"  Formato : {caminho.suffix.upper()}")
    print()

    # Ler
    try:
        linhas, headers = ler_arquivo(caminho)
    except Exception as e:
        print(f"  ❌  Erro ao ler arquivo: {e}")
        sys.exit(1)

    # Verificar colunas obrigatórias
    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in headers]
    if faltando:
        print(f"  ❌  Colunas obrigatórias ausentes: {', '.join(faltando)}")
        print(f"      Colunas encontradas: {', '.join(headers)}")
        print(f"\n  💡  Gere o template com:  python3 scripts/importar_estudantes.py\n")
        sys.exit(1)

    # Filtrar linhas vazias
    linhas = [r for r in linhas if any(
        clean(v) for k, v in r.items() if k == "rm_estudante"
    )]

    print(f"  Linhas encontradas: {len(linhas)}")
    print()

    # Carregar JSON existente (compatível com formato legado que usa "matricula")
    existentes: dict[str, dict] = {}
    if JSON_PATH.is_file():
        with open(JSON_PATH, encoding="utf-8") as f:
            for a in json.load(f):
                # migrar campo legado "matricula" → "rm_estudante"
                if "rm_estudante" not in a and "matricula" in a:
                    a["rm_estudante"] = a["matricula"]
                rm = a.get("rm_estudante") or a.get("matricula", "")
                if rm:
                    existentes[rm] = a

    # Processar
    novos = atualizados = ignorados = 0
    for i, row in enumerate(linhas, start=4):  # começa na linha 4 no XLSX
        aluno = construir_aluno(row, i)
        if aluno is None:
            ignorados += 1
            continue
        rm = aluno["rm_estudante"]
        if rm in existentes:
            existentes[rm] = aluno
            atualizados += 1
        else:
            existentes[rm] = aluno
            novos += 1

    # Ordenar
    from functools import cmp_to_key
    ORDEM_SERIES = ["6º Ano","7º Ano","8º Ano","9º Ano","1º EM","2º EM","3º EM"]

    def cmp_aluno(a, b):
        ia = ORDEM_SERIES.index(a["serie"]) if a["serie"] in ORDEM_SERIES else 99
        ib = ORDEM_SERIES.index(b["serie"]) if b["serie"] in ORDEM_SERIES else 99
        if ia != ib: return ia - ib
        if a["turma"] != b["turma"]: return (a["turma"] > b["turma"]) - (a["turma"] < b["turma"])
        if a["turno"] != b["turno"]: return (a["turno"] > b["turno"]) - (a["turno"] < b["turno"])
        return (a["nome_completo"] > b["nome_completo"]) - (a["nome_completo"] < b["nome_completo"])

    resultado = sorted(existentes.values(), key=cmp_to_key(cmp_aluno))

    # Salvar JSON
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"  {SEP}")
    print(f"  {'✅  Novos':<25}: {novos}")
    print(f"  {'🔄  Atualizados':<25}: {atualizados}")
    print(f"  {'⏭️  Ignorados':<25}: {ignorados}")
    print(f"  {'📋  Total no JSON':<25}: {len(resultado)}")
    print(f"  {'💾  Salvo em':<25}: {JSON_PATH}")
    print(f"  {SEP}")

    # Organizar imagens
    print("\n  Organizando imagens por turma...")
    stats = organizar_imagens(resultado, verbose=True)

    print(f"\n  {SEP}")
    print("  Organização de Imagens:")
    print(f"  {'📁  Pastas criadas':<30}: {len(stats['turmas'])}")
    print(f"  {'🖼️  Imagens copiadas':<30}: {stats['copiadas']}")
    if stats["turmas"]:
        print(f"  {'📂  Pastas':<30}:")
        for t in stats["turmas"]:
            count = len(list((BASE / t).glob("*.jpg"))) if (BASE / t).exists() else 0
            print(f"       • {t}/ ({count} fotos)")
    if stats["nao_encontradas"]:
        print(f"\n  ⚠️  {len(stats['nao_encontradas'])} foto(s) marcadas mas não encontradas em images/:")
        for rm in stats["nao_encontradas"][:10]:
            print(f"       • {rm}.jpg")
        if len(stats["nao_encontradas"]) > 10:
            print(f"       ... e mais {len(stats['nao_encontradas'])-10}")
    print(f"  {SEP}\n")

# ── Gerar templates ────────────────────────────────────────────────────────────
def gerar_templates() -> None:
    print(f"\n{SEP2}")
    print("  Carômetro Escolar — Geração de Templates")
    print(SEP2)

    gerou_algum = False

    if not TMPL_CSV.exists():
        criar_template_csv(TMPL_CSV)
        gerou_algum = True
    else:
        print(f"  ℹ️  template_estudantes.csv já existe (não sobrescrito)")

    if not TMPL_XLSX.exists():
        criar_template_xlsx(TMPL_XLSX)
        gerou_algum = True
    else:
        print(f"  ℹ️  template_estudantes.xlsx já existe (não sobrescrito)")

    if gerou_algum:
        print(f"\n  📌  Próximos passos:")
        print(f"       1. Abra o template e preencha os dados dos estudantes")
        print(f"       2. Coloque as fotos em  images/<rm_estudante>.jpg")
        print(f"       3. Execute:")
        print(f"            python3 scripts/importar_estudantes.py template_estudantes.xlsx")
        print(f"          ou")
        print(f"            python3 scripts/importar_estudantes.py template_estudantes.csv")
    else:
        print(f"\n  💡  Para forçar a recriação, delete os arquivos e execute novamente.")

    print(f"{SEP}\n")

# ── Apenas organizar imagens ───────────────────────────────────────────────────
def apenas_organizar() -> None:
    print(f"\n{SEP2}")
    print("  Carômetro Escolar — Organização de Imagens")
    print(SEP2)

    if not JSON_PATH.is_file():
        print(f"  ❌  {JSON_PATH} não encontrado.")
        print(f"      Execute a importação primeiro.\n")
        sys.exit(1)

    with open(JSON_PATH, encoding="utf-8") as f:
        alunos = json.load(f)

    print(f"  Alunos no JSON: {len(alunos)}")
    print()
    stats = organizar_imagens(alunos)

    print(f"\n  {SEP}")
    print(f"  {'📁  Pastas criadas':<30}: {len(stats['turmas'])}")
    print(f"  {'🖼️  Imagens copiadas':<30}: {stats['copiadas']}")
    if stats["turmas"]:
        for t in stats["turmas"]:
            count = len(list((BASE / t).glob("*.jpg"))) if (BASE / t).exists() else 0
            print(f"       • {t}/ ({count} fotos)")
    print(f"  {SEP}\n")

# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    args = sys.argv[1:]

    if not args:
        gerar_templates()
        return

    if args[0] in ("--organizar-imagens", "--organizar", "-o"):
        apenas_organizar()
        return

    caminho = Path(args[0])
    if not caminho.exists():
        # Tentar relativo à raiz do projeto
        caminho = BASE / args[0]
    if not caminho.exists():
        print(f"\n  ❌  Arquivo não encontrado: {args[0]}")
        print(f"      Gere o template com: python3 scripts/importar_estudantes.py\n")
        sys.exit(1)

    importar(caminho)


if __name__ == "__main__":
    main()
